import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings
import io
import time
import base64
import os

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bluetown CDR Generator",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Logo base64 ───────────────────────────────────────────────────────────────
LOGO_PATH = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
def get_logo_b64():
    try:
        with open(LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

LOGO_B64 = get_logo_b64()

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Space+Grotesk:wght@400;500;600;700&display=swap');

  /* ── Reset & base ── */
  * {{ box-sizing: border-box; }}

  html, body, [data-testid="stAppViewContainer"] {{
    background: #0a0f1e !important;
    font-family: 'Inter', sans-serif;
    color: #e2e8f0;
  }}

  [data-testid="stAppViewContainer"] > .main {{
    background: transparent !important;
    padding: 0 !important;
  }}

  .block-container, [data-testid="stAppViewBlockContainer"] {{
    padding-top: 0 !important;
    padding-bottom: 2rem !important;
    padding-left: 8% !important;   /* <--- Added left breathing room */
    padding-right: 8% !important;  /* <--- Added right breathing room */
    margin-top: 0 !important;
    max-width: 100% !important;
  }}

  /* Hide streamlit chrome */
  #MainMenu, header, footer, [data-testid="stToolbar"], [data-testid="stHeader"] {{
    display: none !important; 
  }}

  /* ── Animated gradient background ── */
  body::before {{
    content: '';
    position: fixed;
    inset: 0;
    background:
      radial-gradient(ellipse 80% 60% at 10% 10%, rgba(82,160,217,0.18) 0%, transparent 60%),
      radial-gradient(ellipse 60% 80% at 90% 90%, rgba(36,99,200,0.14) 0%, transparent 60%),
      radial-gradient(ellipse 50% 50% at 50% 50%, rgba(10,20,50,1) 0%, #0a0f1e 100%);
    z-index: 0;
    pointer-events: none;
    animation: bgPulse 12s ease-in-out infinite alternate;
  }}

  @keyframes bgPulse {{
    0%   {{ opacity: 1; }}
    100% {{ opacity: 0.75; }}
  }}

  /* ── Floating grid overlay ── */
  body::after {{
    content: '';
    position: fixed;
    inset: 0;
    background-image:
      linear-gradient(rgba(82,160,217,0.04) 1px, transparent 1px),
      linear-gradient(90deg, rgba(82,160,217,0.04) 1px, transparent 1px);
    background-size: 48px 48px;
    z-index: 0;
    pointer-events: none;
  }}

  /* ── Navbar ── */
  .bt-nav {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 18px 48px;
    background: rgba(10,15,30,0.7);
    backdrop-filter: blur(20px);
    border-bottom: 1px solid rgba(82,160,217,0.15);
    position: sticky;
    top: 0;
    z-index: 100;
  }}

  .bt-nav-brand {{
    display: flex;
    align-items: center;
    gap: 14px;
  }}

  .bt-nav-logo {{
    width: 48px;
    height: 48px;
    object-fit: contain;
    border-radius: 10px;
    background: white;
    padding: 4px;
  }}

  .bt-nav-title {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 18px;
    font-weight: 700;
    color: #fff;
    letter-spacing: -0.3px;
  }}

  .bt-nav-subtitle {{
    font-size: 11px;
    color: #52a0d9;
    font-weight: 400;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    margin-top: 1px;
  }}

  .bt-nav-badge {{
    background: linear-gradient(135deg, #52a0d9 0%, #2464c8 100%);
    color: white;
    font-size: 11px;
    font-weight: 600;
    padding: 5px 14px;
    border-radius: 20px;
    letter-spacing: 0.3px;
  }}

  /* ── Hero section ── */
  .bt-hero {{
    text-align: center;
    padding: 72px 48px 48px;
    position: relative;
  }}

  .bt-hero-eyebrow {{
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: rgba(82,160,217,0.1);
    border: 1px solid rgba(82,160,217,0.25);
    border-radius: 20px;
    padding: 6px 16px;
    font-size: 12px;
    font-weight: 500;
    color: #52a0d9;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 28px;
    animation: fadeInDown 0.7s ease both;
  }}

  .bt-hero-eyebrow span {{
    width: 6px; height: 6px;
    background: #52a0d9;
    border-radius: 50%;
    animation: blink 2s ease infinite;
  }}

  @keyframes blink {{
    0%, 100% {{ opacity: 1; }}
    50%       {{ opacity: 0.2; }}
  }}

  .bt-hero-h1 {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: clamp(36px, 5vw, 64px);
    font-weight: 700;
    color: #fff;
    line-height: 1.1;
    letter-spacing: -1.5px;
    margin: 0 0 20px;
    animation: fadeInUp 0.7s 0.1s ease both;
  }}

  .bt-hero-h1 .accent {{
    background: linear-gradient(135deg, #52a0d9, #7ec8f5);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
  }}

  .bt-hero-desc {{
    font-size: 18px;
    color: #94a3b8;
    max-width: 540px;
    margin: 0 auto 48px !important; /* Forces the block to center */
    text-align: center !important;  /* Forces the text inside to center */
    line-height: 1.6;
    font-weight: 400;
    animation: fadeInUp 0.7s 0.2s ease both;
  }}

  /* ── Stats bar ── */
  .bt-stats {{
    display: flex;
    justify-content: center;
    gap: 0;
    margin: 0 auto 56px;
    max-width: 640px;
    border: 1px solid rgba(82,160,217,0.18);
    border-radius: 16px;
    overflow: hidden;
    background: rgba(255,255,255,0.02);
    animation: fadeInUp 0.7s 0.3s ease both;
  }}

  .bt-stat {{
    flex: 1;
    text-align: center;
    padding: 20px 16px;
    border-right: 1px solid rgba(82,160,217,0.12);
  }}

  .bt-stat:last-child {{ border-right: none; }}

  .bt-stat-num {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 26px;
    font-weight: 700;
    color: #fff;
    line-height: 1;
  }}

  .bt-stat-label {{
    font-size: 11px;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-top: 4px;
  }}

  /* ── Main card / upload panel ── */
  .bt-panel {{
    max-width: 900px;
    margin: 0 auto 64px;
    padding: 0 24px;
    animation: fadeInUp 0.7s 0.35s ease both;
  }}

  .bt-card {{
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(82,160,217,0.18);
    border-radius: 24px;
    overflow: hidden;
    backdrop-filter: blur(16px);
    box-shadow: 0 24px 64px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.06);
    transition: box-shadow 0.3s ease;
  }}

  .bt-card:hover {{
    box-shadow: 0 32px 80px rgba(0,0,0,0.5), 0 0 0 1px rgba(82,160,217,0.25), inset 0 1px 0 rgba(255,255,255,0.08);
  }}

  .bt-card-header {{
    padding: 28px 36px 20px;
    border-bottom: 1px solid rgba(82,160,217,0.1);
    display: flex;
    align-items: center;
    gap: 14px;
  }}

  .bt-card-icon {{
    width: 42px; height: 42px;
    background: linear-gradient(135deg, #52a0d9, #2464c8);
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 20px;
    flex-shrink: 0;
  }}

  .bt-card-title {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 20px;
    font-weight: 600;
    color: #fff;
  }}

  .bt-card-desc {{
    font-size: 13px;
    color: #64748b;
    margin-top: 2px;
  }}

  .bt-card-body {{
    padding: 32px 36px;
  }}

  /* ── Upload zones ── */
  .bt-upload-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    margin-bottom: 24px;
  }}

  .bt-upload-label {{
    font-size: 12px;
    font-weight: 600;
    color: #52a0d9;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 8px;
  }}

  .bt-upload-hint {{
    font-size: 12px;
    color: #475569;
    margin-top: 6px;
  }}

  /* Override Streamlit file uploader */
  [data-testid="stFileUploader"] {{
    background: rgba(82,160,217,0.04) !important;
    border: 1.5px dashed rgba(82,160,217,0.3) !important;
    border-radius: 14px !important;
    transition: all 0.25s ease !important;
  }}

  [data-testid="stFileUploader"]:hover {{
    border-color: rgba(82,160,217,0.65) !important;
    background: rgba(82,160,217,0.08) !important;
  }}

  [data-testid="stFileUploader"] label {{
    color: #94a3b8 !important;
  }}

  [data-testid="stFileUploaderDropzone"] {{
    background: transparent !important;
    border: none !important;
    padding: 20px !important;
  }}

  [data-testid="stFileUploaderDropzoneInstructions"] {{
    color: #64748b !important;
  }}

  /* ── Text input ── */
  .bt-input-wrap {{ margin-bottom: 28px; }}

  [data-testid="stTextInput"] input {{
    background: rgba(82,160,217,0.06) !important;
    border: 1.5px solid rgba(82,160,217,0.25) !important;
    border-radius: 12px !important;
    color: #e2e8f0 !important;
    font-size: 15px !important;
    padding: 12px 16px !important;
    font-family: 'Inter', sans-serif !important;
    transition: border-color 0.2s ease !important;
  }}

  [data-testid="stTextInput"] input:focus {{
    border-color: #52a0d9 !important;
    box-shadow: 0 0 0 3px rgba(82,160,217,0.15) !important;
    outline: none !important;
  }}

  [data-testid="stTextInput"] label {{
    color: #52a0d9 !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
  }}

  /* ── Generate button ── */
  .bt-btn-wrap {{ text-align: center; }}

  [data-testid="stButton"] > button {{
    background: linear-gradient(135deg, #52a0d9 0%, #2464c8 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 14px !important;
    padding: 16px 56px !important;
    font-size: 16px !important;
    font-weight: 700 !important;
    font-family: 'Space Grotesk', sans-serif !important;
    letter-spacing: 0.3px !important;
    cursor: pointer !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 8px 24px rgba(36,100,200,0.35) !important;
    min-width: 260px !important;
  }}

  [data-testid="stButton"] > button:hover {{
    transform: translateY(-2px) !important;
    box-shadow: 0 14px 36px rgba(36,100,200,0.5) !important;
    background: linear-gradient(135deg, #66b0e8 0%, #3574d8 100%) !important;
  }}

  [data-testid="stButton"] > button:active {{
    transform: translateY(0) !important;
  }}

  /* ── Progress / status ── */
  .bt-status {{
    margin-top: 28px;
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid rgba(82,160,217,0.15);
  }}

  .bt-status-row {{
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 12px 20px;
    background: rgba(82,160,217,0.05);
    border-bottom: 1px solid rgba(82,160,217,0.08);
    animation: slideIn 0.3s ease both;
    font-size: 13.5px;
    color: #cbd5e1;
  }}

  .bt-status-row:last-child {{ border-bottom: none; }}
  .bt-status-row.done {{ color: #4ade80; }}
  .bt-status-row.active {{ color: #52a0d9; }}
  .bt-status-row.error {{ color: #f87171; }}

  .bt-status-icon {{ font-size: 16px; width: 20px; text-align: center; }}

  @keyframes slideIn {{
    from {{ opacity: 0; transform: translateX(-8px); }}
    to   {{ opacity: 1; transform: translateX(0); }}
  }}

  /* ── Download section ── */
  .bt-download-box {{
    background: linear-gradient(135deg, rgba(82,160,217,0.12), rgba(36,100,200,0.08));
    border: 1px solid rgba(82,160,217,0.3);
    border-radius: 16px;
    padding: 28px 32px;
    text-align: center;
    margin-top: 24px;
    animation: fadeInUp 0.5s ease both;
  }}

  .bt-download-title {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 18px;
    font-weight: 700;
    color: #fff;
    margin-bottom: 6px;
  }}

  .bt-download-subtitle {{
    font-size: 13px;
    color: #64748b;
    margin-bottom: 20px;
  }}

  .bt-download-stats {{
    display: flex;
    justify-content: center;
    gap: 24px;
    margin-bottom: 24px;
    flex-wrap: wrap;
  }}

  .bt-dl-stat {{
    text-align: center;
  }}

  .bt-dl-stat-num {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 22px;
    font-weight: 700;
    color: #52a0d9;
  }}

  .bt-dl-stat-lbl {{
    font-size: 11px;
    color: #475569;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-top: 2px;
  }}

  /* Download button */
  [data-testid="stDownloadButton"] > button {{
    background: linear-gradient(135deg, #4ade80 0%, #22c55e 100%) !important;
    color: #052e16 !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 14px 44px !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    font-family: 'Space Grotesk', sans-serif !important;
    cursor: pointer !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 6px 20px rgba(34,197,94,0.35) !important;
  }}

  [data-testid="stDownloadButton"] > button:hover {{
    transform: translateY(-2px) !important;
    box-shadow: 0 12px 30px rgba(34,197,94,0.5) !important;
  }}

  /* ── Error box ── */
  .bt-error {{
    background: rgba(248,113,113,0.08);
    border: 1px solid rgba(248,113,113,0.3);
    border-radius: 14px;
    padding: 20px 24px;
    margin-top: 20px;
    color: #fca5a5;
    font-size: 13.5px;
    animation: fadeInUp 0.4s ease both;
  }}

  /* ── How it works ── */
  .bt-how {{
    max-width: 900px;
    margin: 0 auto 72px;
    padding: 0 24px;
  }}

  .bt-how-title {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 22px;
    font-weight: 700;
    color: #fff;
    margin-bottom: 24px;
    text-align: center;
  }}

  .bt-steps {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 16px;
  }}

  .bt-step {{
    background: rgba(255,255,255,0.025);
    border: 1px solid rgba(82,160,217,0.12);
    border-radius: 16px;
    padding: 24px 20px;
    text-align: center;
    transition: all 0.25s ease;
  }}

  .bt-step:hover {{
    border-color: rgba(82,160,217,0.3);
    background: rgba(82,160,217,0.06);
    transform: translateY(-3px);
  }}

  .bt-step-num {{
    font-family: 'Space Grotesk', sans-serif;
    font-size: 32px;
    font-weight: 800;
    background: linear-gradient(135deg, #52a0d9, #2464c8);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    line-height: 1;
    margin-bottom: 12px;
  }}

  .bt-step-title {{
    font-size: 13px;
    font-weight: 600;
    color: #e2e8f0;
    margin-bottom: 6px;
  }}

  .bt-step-desc {{
    font-size: 12px;
    color: #475569;
    line-height: 1.5;
  }}

  /* ── Footer ── */
  .bt-footer {{
    border-top: 1px solid rgba(82,160,217,0.1);
    padding: 28px 48px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: rgba(10,15,30,0.5);
  }}

  .bt-footer-left {{
    font-size: 12px;
    color: #334155;
  }}

  .bt-footer-right {{
    font-size: 12px;
    color: #334155;
  }}

  /* ── Divider ── */
  .bt-divider {{
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(82,160,217,0.2), transparent);
    margin: 0 auto 56px;
    max-width: 900px;
  }}

  /* ── Animations ── */
  @keyframes fadeInDown {{
    from {{ opacity: 0; transform: translateY(-16px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
  }}

  @keyframes fadeInUp {{
    from {{ opacity: 0; transform: translateY(16px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
  }}

  /* ── Spinner override ── */
  [data-testid="stSpinner"] {{
    color: #52a0d9 !important;
  }}

  /* ── Progress bar ── */
  [data-testid="stProgressBar"] > div > div {{
    background: linear-gradient(90deg, #52a0d9, #2464c8) !important;
    border-radius: 4px !important;
  }}

  [data-testid="stProgressBar"] > div {{
    background: rgba(82,160,217,0.1) !important;
    border-radius: 4px !important;
    height: 6px !important;
  }}

  /* ── Success/info alerts ── */
  [data-testid="stAlert"] {{
    border-radius: 12px !important;
    border: none !important;
  }}

  /* ── Responsive ── */
  @media (max-width: 768px) {{
    .bt-upload-grid {{ grid-template-columns: 1fr; }}
    .bt-steps {{ grid-template-columns: 1fr 1fr; }}
    .bt-nav {{ padding: 14px 20px; }}
    .bt-hero {{ padding: 48px 20px 32px; }}
    .bt-card-body {{ padding: 24px 20px; }}
    .bt-card-header {{ padding: 20px; }}
    .bt-stats {{ flex-direction: column; }}
  }}
</style>
""", unsafe_allow_html=True)


# ── CDR Engine ────────────────────────────────────────────────────────────────

RAW_COLS = [
    "Subscriber ID", "Mobile Number", "Session Start", "Session End",
    "Online Time(Hr.)", "Uploaded MB", "Downloaded MB", "Total MB",
    "IP Address", "MAC Address", "BT Site ID", "AP Name", "Hotspot Name",
    "Circle", "Plan Name"
]

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
PVT_FILL  = PatternFill("solid", fgColor="2E75B6")
PVT_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
DATA_FONT = Font(name="Calibri", size=10)


def auto_fit(ws, max_w=45):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)


def load_mtd(master_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=True, read_only=True)
    mtd_mob = set()
    if "MTD(Mobile No.)" in wb.sheetnames:
        for row in wb["MTD(Mobile No.)"].iter_rows(min_row=2, values_only=True):
            if row[1]: mtd_mob.add(str(row[1]).strip())
    mtd_mac = set()
    if "MTD(MAC)" in wb.sheetnames:
        for row in wb["MTD(MAC)"].iter_rows(min_row=2, values_only=True):
            if row[1]: mtd_mac.add(str(row[1]).strip())
    wb.close()
    return mtd_mob, mtd_mac


def load_and_enrich(input_bytes):
    xl = pd.ExcelFile(io.BytesIO(input_bytes))
    df = pd.read_excel(io.BytesIO(input_bytes), sheet_name=xl.sheet_names[0])
    df.columns = [c.strip() for c in df.columns]
    for col in ["Uploaded MB", "Downloaded MB", "Total MB"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    cdr_date = df["Session Start"].dropna().dt.date.mode()[0]
    df["Plan Name"] = "Free User"
    df["Mint"] = (
        (df["Session End"] - df["Session Start"])
        .dt.total_seconds().div(60).clip(lower=0).fillna(0).astype(int)
    )
    def to_hms(m):
        h, mn = divmod(int(m), 60)
        return f"{h:02d}:{mn:02d}:00"
    df["Online Time(Hr.)"] = df["Mint"].apply(to_hms)
    df["abc"] = df["Mobile Number"].astype(str) + "&" + df["BT Site ID"].astype(str)
    df["efg"] = df["MAC Address"].astype(str)   + "&" + df["BT Site ID"].astype(str)
    df["xyz"] = (df["abc"].iloc[::-1].groupby(df["abc"].iloc[::-1]).cumcount().iloc[::-1] + 1)
    df["hij"] = (df["efg"].iloc[::-1].groupby(df["efg"].iloc[::-1]).cumcount().iloc[::-1] + 1)
    return df, cdr_date


def build_filtered(df, mtd_mob, mtd_mac):
    tu_mask     = ~df.duplicated(subset="abc", keep="last")
    tu_df       = df[tu_mask][RAW_COLS].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    uu_mask     = tu_mask & ~df["abc"].isin(mtd_mob)
    uu_df       = df[uu_mask][RAW_COLS+["abc"]].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    mac_tu_mask = ~df.duplicated(subset="efg", keep="last")
    mac_tu_df   = df[mac_tu_mask][RAW_COLS].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    mac_uu_mask = mac_tu_mask & ~df["efg"].isin(mtd_mac)
    mac_uu_df   = df[mac_uu_mask][RAW_COLS+["efg"]].sort_values(["BT Site ID","Session Start"]).reset_index(drop=True)
    return tu_df, uu_df, mac_tu_df, mac_uu_df


def pivot_cdr(df):
    p = df.groupby("BT Site ID").agg(
        a=("Session Start","count"), b=("Downloaded MB","sum"),
        c=("Uploaded MB","sum"),     d=("Total MB","sum")
    ).reset_index()
    p.columns = ["Row Labels","Count of Session Start","Sum of Downloaded MB","Sum of Uploaded MB","Sum of Total MB"]
    return p.sort_values("Row Labels").reset_index(drop=True)


def pivot_mcdr(df, uu_df, tu_df):
    uu_cnt = uu_df.groupby("BT Site ID").size().reset_index(name="UU").rename(columns={"BT Site ID":"Row Labels"})
    tu_cnt = tu_df.groupby("BT Site ID").size().reset_index(name="TU").rename(columns={"BT Site ID":"Row Labels"})
    p = df.groupby("BT Site ID").agg(
        a=("Session Start","count"), b=("Mint","sum"),
        c=("Uploaded MB","sum"),     d=("Downloaded MB","sum"), e=("Total MB","sum")
    ).reset_index()
    p.columns = ["Row Labels","Count of Session Start","Sum of Mint","Sum of Uploaded MB","Sum of Downloaded MB","Sum of Total MB"]
    p = p.merge(uu_cnt,on="Row Labels",how="left").merge(tu_cnt,on="Row Labels",how="left")
    p[["UU","TU"]] = p[["UU","TU"]].fillna(0).astype(int)
    return p.sort_values("Row Labels").reset_index(drop=True)


def plan_pivot(src_df):
    p = (src_df.groupby(["BT Site ID","Plan Name"]).size()
         .unstack(fill_value=0).reset_index()
         .rename(columns={"BT Site ID":"Row Labels"}))
    p["Grand Total"] = p.drop(columns="Row Labels").sum(axis=1)
    return p.sort_values("Row Labels").reset_index(drop=True)


def write_pivot_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    ws.append([]); ws.append([])
    ws.append(list(df.columns))
    for c in ws[3]:
        c.fill = PVT_FILL; c.font = PVT_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    num_cols = {i+1 for i,col in enumerate(df.columns) if df[col].dtype in ("int64","float64") and col != "Row Labels"}
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
        ws.append(row)
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx in range(1, len(df.columns)+1):
            cell = ws.cell(r_idx, c_idx)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
            if fill: cell.fill = fill
            if c_idx in num_cols and isinstance(cell.value,(int,float)):
                cell.number_format = "#,##0"
    auto_fit(ws); ws.freeze_panes = "A4"


def _style_data_rows(ws, col_names, start_row=2):
    dt_idx  = {i+1 for i,c in enumerate(col_names) if c in {"Session Start","Session End"}}
    int_idx = {i+1 for i,c in enumerate(col_names) if c in {"Uploaded MB","Downloaded MB","Total MB"}}
    for r in range(start_row, ws.max_row+1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in dt_idx:
            cell = ws.cell(r,c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
                cell.alignment = Alignment(horizontal="center")
        for c in int_idx:
            cell = ws.cell(r,c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            try:
                cell.value = int(float(cell.value))
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            except: pass


def write_data_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    _style_data_rows(ws, list(df.columns))
    auto_fit(ws); ws.freeze_panes = "A2"


def write_cdr_date_sheet(wb, name, df, mtd_mob, mtd_mac):
    ws = wb.create_sheet(name)
    headers = RAW_COLS + ["abc","xyz","pqr","Mint","efg","hij","klm"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20
    n    = len(df)
    last = n + 1
    raw_vals = df[RAW_COLS].values.tolist()
    abc_vals = df["abc"].tolist()
    efg_vals = df["efg"].tolist()
    pqr_vals = [k if k in mtd_mob else "#N/A" for k in abc_vals]
    klm_vals = [k if k in mtd_mac else "#N/A" for k in efg_vals]
    for i, rv in enumerate(raw_vals):
        r = i + 2
        ws.append(rv + [
            f'=B{r}&"&"&K{r}',
            f'=COUNTIF(P{r}:P{last+1},P{r})',
            pqr_vals[i],
            f'=INT((D{r}-C{r})*1440)',
            f'=J{r}&"&"&K{r}',
            f'=COUNTIF(T{r}:T{last+1},T{r})',
            klm_vals[i],
        ])
    _style_data_rows(ws, headers)
    for r in range(2, n+2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(16, 23):
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
    auto_fit(ws); ws.freeze_panes = "A2"


def generate_cdr(input_bytes, master_bytes, progress_cb, status_cb):
    status_cb("loading_mtd")
    mtd_mob, mtd_mac = load_mtd(master_bytes)
    progress_cb(15)

    status_cb("loading_input")
    df, cdr_date = load_and_enrich(input_bytes)
    progress_cb(30)

    status_cb("filtering")
    tu_df, uu_df, mac_tu_df, mac_uu_df = build_filtered(df, mtd_mob, mtd_mac)
    progress_cb(45)

    status_cb("pivots")
    p_cdr      = pivot_cdr(df)
    p_mcdr     = pivot_mcdr(df, uu_df, tu_df)
    uup_df     = plan_pivot(uu_df)
    tup_df     = plan_pivot(tu_df)
    mac_uup_df = plan_pivot(mac_uu_df)
    mac_tup_df = plan_pivot(mac_tu_df)
    progress_cb(60)

    status_cb("writing")
    wb = Workbook(); wb.remove(wb.active)
    date_sheet = f"CDR{cdr_date.strftime('%Y-%m-%d')}"

    write_pivot_sheet(wb, "CDR",          p_cdr)
    write_pivot_sheet(wb, "M.CDR",        p_mcdr)
    write_cdr_date_sheet(wb, date_sheet,  df, mtd_mob, mtd_mac)
    write_pivot_sheet(wb, "MAC UUP",      mac_uup_df)
    write_data_sheet(wb,  "MAC UU",       mac_uu_df)
    write_pivot_sheet(wb, "MAC TUP",      mac_tup_df)
    write_data_sheet(wb,  "MAC TU",       mac_tu_df)
    write_pivot_sheet(wb, "UUP",          uup_df)
    write_data_sheet(wb,  "UU",           uu_df)
    write_pivot_sheet(wb, "TUP",          tup_df)
    write_data_sheet(wb,  "TU",           tu_df)
    progress_cb(90)

    status_cb("saving")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    progress_cb(100)

    stats = {
        "date": str(cdr_date),
        "sessions": len(df),
        "bts": df["BT Site ID"].nunique(),
        "tu": len(tu_df),
        "uu": len(uu_df),
        "mac_tu": len(mac_tu_df),
        "mac_uu": len(mac_uu_df),
        "sheets": len(wb.sheetnames),
    }
    return buf.getvalue(), stats


# ── UI Helpers ────────────────────────────────────────────────────────────────

STATUS_MSGS = {
    "loading_mtd":  ("🔍", "Loading MTD lookup tables from Master CDR…"),
    "loading_input":("📥", "Parsing input session data…"),
    "filtering":    ("🔧", "Applying TU / UU / MAC filters…"),
    "pivots":       ("📊", "Building pivot tables…"),
    "writing":      ("✏️",  "Writing 11 sheets to workbook…"),
    "saving":       ("💾", "Finalising and compressing file…"),
}


def status_row(icon, text, state="done"):
    cls = f"bt-status-row {state}"
    st.markdown(f'<div class="{cls}"><span class="bt-status-icon">{icon}</span>{text}</div>', unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────────────────
if "result_bytes" not in st.session_state:
    st.session_state.result_bytes = None
if "result_stats" not in st.session_state:
    st.session_state.result_stats = None
if "log" not in st.session_state:
    st.session_state.log = []
if "generating" not in st.session_state:
    st.session_state.generating = False


# ── RENDER ────────────────────────────────────────────────────────────────────


# ── Navbar ────────────────────────────────────────────────────────────────────
logo_img = f'<img src="data:image/png;base64,{LOGO_B64}" class="bt-nav-logo" alt="Bluetown"/>' if LOGO_B64 else "📡"

st.markdown(f"""
<div class="bt-nav">
  <div class="bt-nav-brand">
    {logo_img}
    <div>
      <div class="bt-nav-title">Bluetown CDR Generator</div>
      <div class="bt-nav-subtitle">Connecting the unconnected</div>
    </div>
  </div>
  <div class="bt-nav-badge">CDR Automation v2.0</div>
</div>
""", unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="bt-hero">
  <div class="bt-hero-eyebrow"><span></span>Automated Report Engine</div>
  <h1 class="bt-hero-h1">Generate your <span class="accent">CDR Final</span><br/>in seconds</h1>
  <div class="bt-hero-desc">Upload your Input sheet and Master CDR — get a perfectly structured, formula-accurate Excel report with all 11 sheets ready to use.</div>
  <div class="bt-stats">
    <div class="bt-stat"><div class="bt-stat-num">11</div><div class="bt-stat-label">Output Sheets</div></div>
    <div class="bt-stat"><div class="bt-stat-num">100%</div><div class="bt-stat-label">Formula Accurate</div></div>
    <div class="bt-stat"><div class="bt-stat-num">4</div><div class="bt-stat-label">Filter Types</div></div>
    <div class="bt-stat"><div class="bt-stat-num">MTD</div><div class="bt-stat-label">Lookup Verified</div></div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Upload Panel ──────────────────────────────────────────────────────────────
st.markdown('<div class="bt-panel">', unsafe_allow_html=True)
st.markdown("""
<div class="bt-card">
  <div class="bt-card-header">
    <div class="bt-card-icon">📤</div>
    <div>
      <div class="bt-card-title">Upload Files & Generate</div>
      <div class="bt-card-desc">Provide both source files and your desired output name</div>
    </div>
  </div>
""", unsafe_allow_html=True)

st.markdown('<div class="bt-card-body">', unsafe_allow_html=True)

# Upload grid
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="bt-upload-label">📋 Input Session Data</div>', unsafe_allow_html=True)
    input_file = st.file_uploader(
        "input_cdr",
        type=["xlsx"],
        label_visibility="collapsed",
        key="input_upload"
    )
    if input_file:
        st.markdown(f'<div class="bt-upload-hint">✅ {input_file.name} ({input_file.size/1024:.0f} KB)</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="bt-upload-hint">Drag & drop input.xlsx here</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="bt-upload-label">📚 Master CDR Reference</div>', unsafe_allow_html=True)
    master_file = st.file_uploader(
        "master_cdr",
        type=["xlsx"],
        label_visibility="collapsed",
        key="master_upload"
    )
    if master_file:
        st.markdown(f'<div class="bt-upload-hint">✅ {master_file.name} ({master_file.size/1024:.0f} KB)</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="bt-upload-hint">Drag & drop Master CDR here</div>', unsafe_allow_html=True)

# Output filename
st.markdown('<div class="bt-input-wrap" style="margin-top:20px;">', unsafe_allow_html=True)
output_name = st.text_input(
    "OUTPUT FILE NAME",
    value="CDR_Final_Output",
    placeholder="e.g. CDR2026-06-07_final",
    help="The .xlsx extension will be added automatically"
)

# Generate button
st.markdown('<div class="bt-btn-wrap">', unsafe_allow_html=True)
ready = input_file is not None and master_file is not None
generate_clicked = st.button(
    "⚡  Generate CDR Final" if ready else "⬆  Upload Both Files to Continue",
    disabled=not ready,
    use_container_width=False
)
st.markdown('</div>', unsafe_allow_html=True)

# ── Generate logic ────────────────────────────────────────────────────────────
if generate_clicked and ready:
    st.session_state.result_bytes = None
    st.session_state.result_stats = None
    st.session_state.log = []

    input_bytes  = input_file.read()
    master_bytes = master_file.read()

    progress_bar  = st.progress(0)
    status_container = st.container()

    completed_steps = []

    def progress_cb(pct):
        progress_bar.progress(pct)

    def status_cb(key):
        icon, msg = STATUS_MSGS[key]
        completed_steps.append((icon, msg))
        with status_container:
            st.markdown('<div class="bt-status">', unsafe_allow_html=True)
            for i, (ic, tx) in enumerate(completed_steps):
                is_last = i == len(completed_steps) - 1
                state = "active" if is_last else "done"
                tick  = ic if is_last else "✅"
                st.markdown(
                    f'<div class="bt-status-row {state}"><span class="bt-status-icon">{tick}</span>{tx}</div>',
                    unsafe_allow_html=True
                )
            st.markdown('</div>', unsafe_allow_html=True)

    try:
        result_bytes, stats = generate_cdr(
            input_bytes, master_bytes, progress_cb, status_cb
        )
        st.session_state.result_bytes = result_bytes
        st.session_state.result_stats = stats

        # Final status update — all done
        with status_container:
            st.markdown('<div class="bt-status">', unsafe_allow_html=True)
            for ic, tx in completed_steps:
                st.markdown(f'<div class="bt-status-row done"><span class="bt-status-icon">✅</span>{tx}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    except Exception as e:
        st.markdown(f'<div class="bt-error">❌ <strong>Error:</strong> {str(e)}</div>', unsafe_allow_html=True)

# ── Download section ──────────────────────────────────────────────────────────
if st.session_state.result_bytes and st.session_state.result_stats:
    stats = st.session_state.result_stats
    fname = (output_name.strip() or "CDR_Final_Output")
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    st.markdown(f"""
    <div class="bt-download-box">
      <div class="bt-download-title">✅ CDR Final Ready</div>
      <div class="bt-download-subtitle">Generated from {stats['sessions']:,} sessions across {stats['bts']:,} BTS sites</div>
      <div class="bt-download-stats">
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['sessions']:,}</div>
          <div class="bt-dl-stat-lbl">Total Sessions</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['bts']:,}</div>
          <div class="bt-dl-stat-lbl">BTS Sites</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['tu']:,}</div>
          <div class="bt-dl-stat-lbl">TU Rows</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['uu']:,}</div>
          <div class="bt-dl-stat-lbl">UU Rows</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['mac_tu']:,}</div>
          <div class="bt-dl-stat-lbl">MAC TU</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['mac_uu']:,}</div>
          <div class="bt-dl-stat-lbl">MAC UU</div>
        </div>
        <div class="bt-dl-stat">
          <div class="bt-dl-stat-num">{stats['sheets']}</div>
          <div class="bt-dl-stat-lbl">Sheets</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_dl1, col_dl2, col_dl3 = st.columns([1, 2, 1])
    with col_dl2:
        st.download_button(
            label="⬇  Download CDR Final (.xlsx)",
            data=st.session_state.result_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.markdown('</div></div>', unsafe_allow_html=True)  # close card-body + card
st.markdown('</div>', unsafe_allow_html=True)  # close panel

# ── Divider ───────────────────────────────────────────────────────────────────
st.markdown('<div class="bt-divider"></div>', unsafe_allow_html=True)

# ── How it works ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="bt-how">
  <div class="bt-how-title">How it works</div>
  <div class="bt-steps">
    <div class="bt-step">
      <div class="bt-step-num">01</div>
      <div class="bt-step-title">Upload Files</div>
      <div class="bt-step-desc">Provide your daily Input CDR and the Master CDR reference workbook</div>
    </div>
    <div class="bt-step">
      <div class="bt-step-num">02</div>
      <div class="bt-step-title">MTD Lookup</div>
      <div class="bt-step-desc">Engine loads MTD Mobile & MAC keys to identify new vs returning users</div>
    </div>
    <div class="bt-step">
      <div class="bt-step-num">03</div>
      <div class="bt-step-title">Filter & Enrich</div>
      <div class="bt-step-desc">Applies TU, UU, MAC TU, MAC UU filters with suffix COUNTIF logic</div>
    </div>
    <div class="bt-step">
      <div class="bt-step-num">04</div>
      <div class="bt-step-title">Download</div>
      <div class="bt-step-desc">Get a 11-sheet formatted Excel file with all pivots and formulas intact</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="bt-footer">
  <div class="bt-footer-left">© 2026 Bluetown · Connecting the unconnected</div>
  <div class="bt-footer-right">CDR Automation Engine v2.0 · Built for Bluetown Operations</div>
</div>
""", unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)  # close bt-page
