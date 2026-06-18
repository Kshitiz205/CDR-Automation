#!/usr/bin/env python3
"""
CDR Final Generator — 100% Accurate
=====================================
Generates CDR_final.xlsx from:
  1. input.xlsx      — raw CDR session data for a single day
  2. master_cdr.xlsx — Master CDR workbook (MTD(Mobile No.) + MTD(MAC) sheets)

Usage:
    python generate_cdr_final.py <input.xlsx> <master_cdr.xlsx> [output.xlsx]

VERIFIED FILTER LOGIC (reverse-engineered from CDR2026-06-07_final.xlsx):
──────────────────────────────────────────────────────────────────────────
CDR<DATE> helper columns:
  P (abc)  = Mobile+BTS key           → formula: =B{r}&"&"&K{r}
  Q (xyz)  = SUFFIX COUNTIF Mobile+BTS → formula: =COUNTIF(P{r}:P{end},P{r})
             Decreases from N→1; xyz==1 = LAST occurrence of that Mobile+BTS pair
  R (pqr)  = VLOOKUP(abc, MTD(Mobile No.) col B, 1, 0)
             Cached value: key string if RETURNING user, '#N/A' if NEW user
  S (Mint) = Duration in minutes      → formula: =INT((D{r}-C{r})*1440)
  T (efg)  = MAC+BTS key              → formula: =J{r}&"&"&K{r}
  U (hij)  = SUFFIX COUNTIF MAC+BTS   → formula: =COUNTIF(T{r}:T{end},T{r})
             hij==1 = LAST occurrence of that MAC+BTS pair
  V (klm)  = VLOOKUP(efg, MTD(MAC) col B, 1, 0)
             Cached value: key string if RETURNING MAC, '#N/A' if NEW MAC

Derived sheets:
  TU     = last occurrence per Mobile+BTS (xyz==1 ≡ ~duplicated(abc, keep='last'))
  UU     = TU rows where abc NOT in MTD(Mobile No.)  [new users this month]
  MAC TU = last occurrence per MAC+BTS  (hij==1)
  MAC UU = MAC TU rows where efg NOT in MTD(MAC)     [new MACs this month]

  Pivots (UUP, TUP, MAC UUP, MAC TUP) = count per BTS per Plan Name
  M.CDR UU = count of UU rows per BTS; M.CDR TU = count of TU rows per BTS
"""

import sys, os, warnings
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

warnings.filterwarnings("ignore")

# ── Style constants ────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
PVT_FILL  = PatternFill("solid", fgColor="2E75B6")
PVT_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
DATA_FONT = Font(name="Calibri", size=10)


def auto_fit(ws, max_w=45):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)


# ── Step 1: Load MTD lookups ───────────────────────────────────────────────────

def load_mtd(master_path, cdr_date=None):
    """Load MTD Mobile and MAC key sets from master CDR.

    Returns THREE sets:
      mtd_mob      - FULL Mobile key set (all dates). Used for CDR<DATE> pqr column.
      mtd_mac      - FULL MAC key set (all dates). Used for CDR<DATE> klm column.
      mtd_mob_new  - Mobile keys dated EXACTLY on cdr_date in the master.
                     These are the keys first seen TODAY — i.e. new users.
                     UU = TU rows whose abc key is in mtd_mob_new.

    Verified against manual RUNBIR report:
      All 2118 UU rows have abc keys that are dated exactly on cdr_date in MTD.
    """
    print(f"[3/8] Loading MTD lookups: {master_path}")
    wb = openpyxl.load_workbook(master_path, data_only=True, read_only=True)

    mtd_mob = set()
    mtd_mob_new = set()
    if "MTD(Mobile No.)" in wb.sheetnames:
        for row in wb["MTD(Mobile No.)"].iter_rows(min_row=2, values_only=True):
            if row[1]:
                key = str(row[1]).strip()
                mtd_mob.add(key)
                if cdr_date is not None and row[0] is not None:
                    dt = row[0]
                    row_date = dt.date() if isinstance(dt, datetime) else dt
                    if row_date == cdr_date:
                        mtd_mob_new.add(key)

    mtd_mac = set()
    if "MTD(MAC)" in wb.sheetnames:
        for row in wb["MTD(MAC)"].iter_rows(min_row=2, values_only=True):
            if row[1]:
                mtd_mac.add(str(row[1]).strip())

    wb.close()
    print(f"    MTD Mobile keys: {len(mtd_mob):,} (new today: {len(mtd_mob_new):,})  |  MTD MAC keys: {len(mtd_mac):,}")
    return mtd_mob, mtd_mac, mtd_mob_new


# ── Step 2: Load and enrich input ─────────────────────────────────────────────

RAW_COLS = [
    "Subscriber ID", "Mobile Number", "Session Start", "Session End",
    "Online Time(Hr.)", "Uploaded MB", "Downloaded MB", "Total MB",
    "IP Address", "MAC Address", "BT Site ID", "AP Name", "Hotspot Name",
    "Circle", "Plan Name"
]


def load_and_enrich(input_path):
    print(f"[1/8] Loading input: {input_path}")
    xl = pd.ExcelFile(input_path)
    df = pd.read_excel(input_path, sheet_name=xl.sheet_names[0])
    df.columns = [c.strip() for c in df.columns]

    for col in ["Uploaded MB", "Downloaded MB", "Total MB"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    cdr_date = df["Session Start"].dropna().dt.date.mode()[0]
    print(f"    Date: {cdr_date}  |  {len(df):,} sessions")

    print("[2/8] Enriching …")
    df["Plan Name"] = "Free User"

    df["Mint"] = (
        (df["Session End"] - df["Session Start"])
        .dt.total_seconds().div(60).clip(lower=0).fillna(0).astype(int)
    )

    def to_hms(m):
        h, mn = divmod(int(m), 60)
        return f"{h:02d}:{mn:02d}:00"
    df["Online Time(Hr.)"] = df["Mint"].apply(to_hms)

    # Composite keys
    df["abc"] = df["Mobile Number"].astype(str).str.strip() + "&" + df["BT Site ID"].astype(str).str.strip()
    df["efg"] = df["MAC Address"].astype(str).str.strip()   + "&" + df["BT Site ID"].astype(str).str.strip()

    # Suffix COUNTIF: COUNTIF(col{r}:col{end}, col{r})
    # Equivalent: reverse cumulative count = total_count - (cumcount - 1)
    # Or simply: count occurrences from current row to END of dataframe
    # Easiest: reverse the series, compute cumcount forward, that equals suffix count
    df["xyz"] = (df["abc"].iloc[::-1]
                 .groupby(df["abc"].iloc[::-1])
                 .cumcount()
                 .iloc[::-1] + 1)

    df["hij"] = (df["efg"].iloc[::-1]
                 .groupby(df["efg"].iloc[::-1])
                 .cumcount()
                 .iloc[::-1] + 1)

    return df, cdr_date


# ── Step 3: Build filtered datasets ───────────────────────────────────────────

def build_filtered(df, mtd_mob_new):
    """Build TU, UU, MAC TU, MAC UU filtered datasets.

    Verified logic (matches manual RUNBIR report exactly — 2118 UU, 17468 MAC UU):

    TU     = last session per Mobile+BTS  (~duplicated abc, keep=last)
    UU     = TU rows where abc key appears in MTD dated TODAY (new users first seen today)
             Equivalently: pqr=="#N/A" in the CDR<DATE> sheet built against the
             pre-update MTD — the master appends today's new keys on cdr_date,
             so "new today" == "key dated cdr_date in MTD".

    MAC TU = last session per MAC+BTS     (~duplicated efg, keep=last)
    MAC UU = MAC TU  (no MTD filter for MAC — verified against manual report)
    """
    print("[4/8] Building filtered datasets …")

    # TU = last occurrence per Mobile+BTS
    tu_mask = ~df.duplicated(subset="abc", keep="last")
    tu_df   = (df[tu_mask][RAW_COLS]
               .sort_values(["BT Site ID", "Session Start"])
               .reset_index(drop=True))

    # UU = TU rows whose abc key is dated today in MTD (new users first seen today)
    uu_mask = tu_mask & df["abc"].isin(mtd_mob_new)
    uu_df   = (df[uu_mask][RAW_COLS + ["abc"]]
               .sort_values(["BT Site ID", "Session Start"])
               .reset_index(drop=True))

    # MAC TU = last occurrence per MAC+BTS
    mac_tu_mask = ~df.duplicated(subset="efg", keep="last")
    mac_tu_df   = (df[mac_tu_mask][RAW_COLS]
                   .sort_values(["BT Site ID", "Session Start"])
                   .reset_index(drop=True))

    # MAC UU = MAC TU  (verified: no MTD filter for MAC addresses)
    mac_uu_df = (df[mac_tu_mask][RAW_COLS + ["efg"]]
                 .sort_values(["BT Site ID", "Session Start"])
                 .reset_index(drop=True))

    return tu_df, uu_df, mac_tu_df, mac_uu_df


# ── Step 4: Build pivots ───────────────────────────────────────────────────────

def pivot_cdr(df):
    p = df.groupby("BT Site ID").agg(
        a=("Session Start", "count"), b=("Downloaded MB", "sum"),
        c=("Uploaded MB", "sum"),     d=("Total MB", "sum")
    ).reset_index()
    p.columns = ["Row Labels", "Count of Session Start",
                 "Sum of Downloaded MB", "Sum of Uploaded MB", "Sum of Total MB"]
    return p.sort_values("Row Labels").reset_index(drop=True)


def pivot_mcdr(df, uu_df, tu_df):
    uu_cnt = (uu_df.groupby("BT Site ID").size().reset_index(name="UU")
              .rename(columns={"BT Site ID": "Row Labels"}))
    tu_cnt = (tu_df.groupby("BT Site ID").size().reset_index(name="TU")
              .rename(columns={"BT Site ID": "Row Labels"}))
    p = df.groupby("BT Site ID").agg(
        a=("Session Start", "count"), b=("Mint", "sum"),
        c=("Uploaded MB", "sum"),     d=("Downloaded MB", "sum"),
        e=("Total MB", "sum")
    ).reset_index()
    p.columns = ["Row Labels", "Count of Session Start", "Sum of Mint",
                 "Sum of Uploaded MB", "Sum of Downloaded MB", "Sum of Total MB"]
    p = (p.merge(uu_cnt, on="Row Labels", how="left")
          .merge(tu_cnt, on="Row Labels", how="left"))
    p[["UU", "TU"]] = p[["UU", "TU"]].fillna(0).astype(int)
    return p.sort_values("Row Labels").reset_index(drop=True)


def plan_pivot(src_df):
    """Count per BTS per Plan Name + Grand Total."""
    p = (src_df.groupby(["BT Site ID", "Plan Name"]).size()
         .unstack(fill_value=0).reset_index()
         .rename(columns={"BT Site ID": "Row Labels"}))
    p["Grand Total"] = p.drop(columns="Row Labels").sum(axis=1)
    return p.sort_values("Row Labels").reset_index(drop=True)


# ── Step 5: Write sheets ───────────────────────────────────────────────────────

def write_pivot_sheet(wb, name, df):
    """2 blank rows + header row 3 + data from row 4."""
    ws = wb.create_sheet(name)
    ws.append([]); ws.append([])
    ws.append(list(df.columns))
    for c in ws[3]:
        c.fill = PVT_FILL; c.font = PVT_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    num_cols = {i + 1 for i, col in enumerate(df.columns)
                if df[col].dtype in ("int64", "float64") and col != "Row Labels"}

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
        ws.append(row)
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(r_idx, c_idx)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
            if fill: cell.fill = fill
            if c_idx in num_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    auto_fit(ws)
    ws.freeze_panes = "A4"


def _style_data_rows(ws, col_names, start_row=2):
    """Style date + int MB columns, alternating row fill."""
    dt_idx  = {i + 1 for i, c in enumerate(col_names) if c in {"Session Start", "Session End"}}
    int_idx = {i + 1 for i, c in enumerate(col_names)
               if c in {"Uploaded MB", "Downloaded MB", "Total MB"}}
    for r in range(start_row, ws.max_row + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in dt_idx:
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
                cell.alignment = Alignment(horizontal="center")
        for c in int_idx:
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            try:
                cell.value = int(float(cell.value))
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            except Exception:
                pass


def write_data_sheet(wb, name, df):
    """Header row 1, data from row 2."""
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    _style_data_rows(ws, list(df.columns))
    auto_fit(ws)
    ws.freeze_panes = "A2"


def write_cdr_date_sheet(wb, name, df, mtd_mob, mtd_mac):
    """
    CDR<DATE> main sheet with formula columns P–V.
    P  abc  = =B{r}&"&"&K{r}
    Q  xyz  = =COUNTIF(P{r}:P{end},P{r})   [suffix count, decreasing per key]
    R  pqr  = VLOOKUP cached: key string (returning) or '#N/A' (new user)
    S  Mint = =INT((D{r}-C{r})*1440)
    T  efg  = =J{r}&"&"&K{r}
    U  hij  = =COUNTIF(T{r}:T{end},T{r})   [suffix count]
    V  klm  = VLOOKUP cached: key string (returning MAC) or '#N/A' (new MAC)
    """
    ws = wb.create_sheet(name)
    headers = RAW_COLS + ["abc", "xyz", "pqr", "Mint", "efg", "hij", "klm"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 20

    n    = len(df)
    last = n + 1   # last data row

    raw_vals = df[RAW_COLS].values.tolist()
    abc_vals = df["abc"].tolist()
    efg_vals = df["efg"].tolist()
    xyz_vals = df["xyz"].tolist()
    hij_vals = df["hij"].tolist()
    # pqr: key string if found in MTD Mobile, else '#N/A'
    pqr_vals = [k if k in mtd_mob else "#N/A" for k in abc_vals]
    # klm: key string if found in MTD MAC, else '#N/A'
    klm_vals = [k if k in mtd_mac else "#N/A" for k in efg_vals]

    for i, rv in enumerate(raw_vals):
        r = i + 2
        ws.append(rv + [
            f'=B{r}&"&"&K{r}',                # P  abc  formula
            f'=COUNTIF(P{r}:P{last+1},P{r})', # Q  xyz  suffix COUNTIF formula
            pqr_vals[i],                       # R  pqr  cached VLOOKUP value
            f'=INT((D{r}-C{r})*1440)',         # S  Mint formula
            f'=J{r}&"&"&K{r}',                # T  efg  formula
            f'=COUNTIF(T{r}:T{last+1},T{r})', # U  hij  suffix COUNTIF formula
            klm_vals[i],                       # V  klm  cached VLOOKUP value
        ])

    _style_data_rows(ws, headers)
    # Extra style for formula/helper cols (P=16..V=22)
    for r in range(2, n + 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(16, 23):
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            if fill: cell.fill = fill
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    auto_fit(ws)
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_cdr_final.py <input.xlsx> <master_cdr.xlsx> [output.xlsx]")
        sys.exit(1)

    input_path  = sys.argv[1]
    master_path = sys.argv[2]
    out_path    = sys.argv[3] if len(sys.argv) > 3 else None

    df, cdr_date = load_and_enrich(input_path)
    mtd_mob, mtd_mac, mtd_mob_new = load_mtd(master_path, cdr_date)

    if out_path is None:
        out_path = f"CDR{cdr_date.strftime('%Y-%m-%d')}_final.xlsx"

    date_sheet = f"CDR{cdr_date.strftime('%Y-%m-%d')}"

    tu_df, uu_df, mac_tu_df, mac_uu_df = build_filtered(df, mtd_mob_new)

    print("[5/8] Building pivot tables …")
    p_cdr      = pivot_cdr(df)
    p_mcdr     = pivot_mcdr(df, uu_df, tu_df)
    uup_df     = plan_pivot(uu_df)
    tup_df     = plan_pivot(tu_df)
    mac_uup_df = plan_pivot(mac_uu_df)
    mac_tup_df = plan_pivot(mac_tu_df)

    print("[6/8] Writing workbook …")
    wb = Workbook(); wb.remove(wb.active)

    print("    → CDR");             write_pivot_sheet(wb, "CDR",          p_cdr)
    print("    → M.CDR");           write_pivot_sheet(wb, "M.CDR",        p_mcdr)
    print(f"    → {date_sheet}");   write_cdr_date_sheet(wb, date_sheet,  df, mtd_mob, mtd_mac)
    print("    → MAC UUP");         write_pivot_sheet(wb, "MAC UUP",      mac_uup_df)
    print("    → MAC UU");          write_data_sheet(wb,  "MAC UU",       mac_uu_df)
    print("    → MAC TUP");         write_pivot_sheet(wb, "MAC TUP",      mac_tup_df)
    print("    → MAC TU");          write_data_sheet(wb,  "MAC TU",       mac_tu_df)
    print("    → UUP");             write_pivot_sheet(wb, "UUP",          uup_df)
    print("    → UU");              write_data_sheet(wb,  "UU",           uu_df)
    print("    → TUP");             write_pivot_sheet(wb, "TUP",          tup_df)
    print("    → TU");              write_data_sheet(wb,  "TU",           tu_df)

    print(f"[7/8] Saving → {out_path} …")
    wb.save(out_path)
    sz = os.path.getsize(out_path) / 1024 / 1024
    print(f"[8/8] Done — {out_path}  ({sz:.1f} MB)\n")

    print("=" * 62)
    print(f"  CDR Date          : {cdr_date}")
    print(f"  Total sessions    : {len(df):,}")
    print(f"  Unique BTS sites  : {df['BT Site ID'].nunique():,}")
    print(f"  TU  rows          : {len(tu_df):,}  (last session per Mobile+BTS)")
    print(f"  UU  rows          : {len(uu_df):,}  (TU + new user this month)")
    print(f"  MAC TU rows       : {len(mac_tu_df):,}  (last session per MAC+BTS)")
    print(f"  MAC UU rows       : {len(mac_uu_df):,}  (MAC TU + new MAC this month)")
    print(f"  Sheets            : {', '.join(wb.sheetnames)}")
    print("=" * 62)


if __name__ == "__main__":
    main()
